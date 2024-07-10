import pywencai
import os
import datetime
import pandas as res
import openpyxl
from openpyxl import load_workbook

# 获取关键字列表
keywords = os.environ.get("KEYWORD", "").split(',')
today = datetime.date.today()
tomorrow = today + datetime.timedelta(days=1)
t2 = today.strftime('%m月%d日')

# 遍历关键字列表进行查询和保存结果
for keyword in keywords:
  filename = f'{keyword}-{tomorrow}.xlsx'
  keyword2 = t2 + ' ' + keyword
  res = pywencai.get(query=keyword2, loop=True)
  columns_to_select = [0, 1, 5, 10]
  # 使用iloc来截取列
  selected_res = res.iloc[:, columns_to_select]
  if res is None:
    raise ValueError("Did not receive any data. The result is None.")
  #selected_res.to_excel(filename, index=False)  # index=False表示不写入行索引

  workbook_path = 'module.xlsx'
  workbook = load_workbook(workbook_path)
  workbook.calc_on_load = True
  sheet= workbook.active
  #确定行数
  res_rows = len(selected_res)
  sheet_rows = sheet.max_row
  for i in range(res_rows):
    sheet_row = i + 1
    if sheet_row <= sheet_rows:
      sheet[f'B{i+2}'] = selected_res.iloc[i,0]
      sheet[f'F{i+2}'] = selected_res.iloc[i,3]

  #workbook.calculate()
  output_path = f'季报汇总-{tomorrow}.xlsx'
  workbook.save(output_path)


  
